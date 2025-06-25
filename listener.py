# from fast_plate_ocr import ONNXPlateRecognizer
# import cv2
# import time
# import os
# import re
# import sys
# import json

# def get_highest_confidence_images():
#     image_folder_path = './plate_crops'
#     if not os.path.exists(image_folder_path):
#         print(f"Error: Directory {image_folder_path} does not exist")
#         return []
        
#     files = [f for f in os.listdir(image_folder_path) if os.path.isfile(os.path.join(image_folder_path, f))]
    
#     plate_groups = {}
#     for file in files:
#         match = re.match(r'plate_(\d+)_\d+_(\d+\.\d+)\.jpg', file)
#         if match:
#             plate_id = match.group(1)
#             confidence = float(match.group(2))
            
#             if plate_id not in plate_groups or confidence > plate_groups[plate_id][1]:
#                 plate_groups[plate_id] = (file, confidence)
    
#     highest_conf_images = [os.path.join(image_folder_path, info[0]) for info in plate_groups.values()]
    
#     for file in files:
#         file_path = os.path.join(image_folder_path, file)
#         if file_path not in highest_conf_images:
#             try:
#                 os.remove(file_path)
#                 print(f"Deleted: {file_path}")
#             except Exception as e:
#                 print(f"Error deleting {file_path}: {e}")
    
#     return highest_conf_images

# def process_plate_image(image_path):
#     image = cv2.imread(image_path)
#     if image is None:
#         print("Error: Could not read image")
#         return None
    
#     gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

#     try:
#         m = ONNXPlateRecognizer('global-plates-mobile-vit-v2-model', providers=['CPUExecutionProvider'])
#     except Exception as e:
#         print(f"Error initializing ONNXPlateRecognizer: {e}")
#         return None

#     start_time = time.time()
#     results = m.run(gray_image)
#     onnx_time = time.time() - start_time
    
#     print(f"ONNXPlateRecognizer result: {results}")
    
#     if not results:
#         print("No license plate detected")
#         return None

#     plate_text = results[0] if isinstance(results, list) else results
#     plate_text = re.sub(r'[^a-zA-Z0-9]', '', plate_text).upper()
 
#     letters = re.sub(r'[^A-Z]', '', plate_text)
#     numbers = re.sub(r'[^0-9]', '', plate_text)

#     numbers = numbers[:4]

#     formatted_plate = f"{letters}-{numbers}" if letters and numbers else plate_text

#     confidence = 1.0
#     if hasattr(m, 'confidence'):
#         confidence = float(m.confidence)
    
#     print(f"OCR Confidence: {confidence:.4f}")

#     result = {
#         'plate_text': formatted_plate,
#         'confidence': confidence,
#         'processing_time': onnx_time
#     }

#     if len(re.findall(r'[A-Z]', formatted_plate)) >= 2:
#         save_plate_to_file(formatted_plate, confidence)
#     else:
#         print(f"Skipping license plate '{formatted_plate}' as it does not contain at least two alphabet characters")
    
#     return result

# def save_plate_to_file(plate_text, confidence=None, filename="license_plates.txt"):
#     timestamp = time.strftime("%Y-%m-%d %H:%M:%S")

#     existing_entries = []
#     if os.path.exists(filename):
#         with open(filename, 'r') as file:
#             existing_entries = file.readlines()

#     filtered_entries = []
#     seen_plates = set()
#     for entry in existing_entries:
#         entry_parts = entry.strip().split(' - ')
#         if len(entry_parts) > 1:
#             existing_plate = entry_parts[1].split(' ')[0]
#             if existing_plate not in seen_plates:
#                 seen_plates.add(existing_plate)
#                 filtered_entries.append(entry)

#     confidence_str = f" (Confidence: {confidence:.4f})" if confidence is not None else ""
#     new_entry = f"{timestamp} - {plate_text}{confidence_str}\n"

#     if plate_text not in seen_plates or (confidence and confidence > 0):
#         filtered_entries.insert(0, new_entry)

#     with open(filename, 'w') as file:
#         file.writelines(filtered_entries[:50])
    
#     print(f"License plate '{plate_text}' saved to {filename}{confidence_str}")

# if __name__ == "__main__":
#     if len(sys.argv) > 1:
#         image_path = sys.argv[1]
#         result = process_plate_image(image_path)
#         if result:
#             print(f"OCR_RESULT:{json.dumps(result)}")
#     else:
#         highest_conf_images = get_highest_confidence_images()
        
#         if not highest_conf_images:
#             print("No images found in the plate_crops directory")
#             sys.exit(1)
            
#         for image_path in highest_conf_images:
#             print(f"Processing: {image_path}")
#             result = process_plate_image(image_path)
#             if result:
#                 print(f"OCR_RESULT:{json.dumps(result)}")

from fast_plate_ocr import ONNXPlateRecognizer
import cv2
import time
import os
import re
import sys
import json
import base64
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import tempfile

def get_highest_confidence_images():
    image_folder_path = './plate_crops'
    if not os.path.exists(image_folder_path):
        return []
    files = [f for f in os.listdir(image_folder_path) if os.path.isfile(os.path.join(image_folder_path, f))]
    plate_groups = {}
    for file in files:
        match = re.match(r'plate_(\d+)_\d+_(\d+\.\d+)\.jpg', file)
        if match:
            plate_id = match.group(1)
            confidence = float(match.group(2))
            if plate_id not in plate_groups or confidence > plate_groups[plate_id][1]:
                plate_groups[plate_id] = (file, confidence)
    highest_conf_images = [os.path.join(image_folder_path, info[0]) for info in plate_groups.values()]
    for file in files:
        file_path = os.path.join(image_folder_path, file)
        if file_path not in highest_conf_images:
            try:
                os.remove(file_path)
            except:
                pass
    return highest_conf_images

def save_plate_to_excel(entry, filename="license_plates.xlsx"):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['timestamp', 'plate_text', 'processing_time', 'image_filename', 'plate_image'])

    existing_plates = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        plate, ts = row[1], row[0]
        if plate not in existing_plates or ts > existing_plates[plate]:
            existing_plates[plate] = ts

    if entry['plate_text'] in existing_plates and entry['timestamp'] <= existing_plates[entry['plate_text']]:
        wb.save(filename)
        return

    img_path = None
    with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp:
        tmp.write(base64.b64decode(entry['plate_image']))
        img_path = tmp.name

    ws.append([entry['timestamp'], entry['plate_text'], entry['processing_time'], entry['image_filename'], None])
    img = XLImage(img_path)
    img.width = 120
    img.height = 40
    img_cell = f"E{ws.max_row}"
    ws.add_image(img, img_cell)

    wb.save(filename)
    os.remove(img_path)

def encode_image_to_base64(image):
    _, buffer = cv2.imencode('.jpg', image)
    return base64.b64encode(buffer).decode('utf-8')

def process_plate_image(image_path, recognizer):
    image = cv2.imread(image_path)
    if image is None:
        return None
    gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    start_time = time.time()
    results = recognizer.run(gray_image)
    onnx_time = time.time() - start_time
    if not results:
        return None
    plate_text = results[0] if isinstance(results, list) else results
    plate_text = re.sub(r'[^a-zA-Z0-9]', '', plate_text).upper()
    letters = re.sub(r'[^A-Z]', '', plate_text)
    numbers = re.sub(r'[^0-9]', '', plate_text)
    numbers = numbers[:4]
    formatted_plate = f"{letters}-{numbers}" if letters and numbers else plate_text
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
    if len(re.findall(r'[A-Z]', formatted_plate)) >= 2:
        entry = {
            'timestamp': timestamp,
            'plate_text': formatted_plate,
            'processing_time': round(onnx_time, 4),
            'image_filename': os.path.basename(image_path),
            'plate_image': encode_image_to_base64(image)
        }
        save_plate_to_excel(entry)
        return entry
    return None

if __name__ == "__main__":
    try:
        recognizer = ONNXPlateRecognizer('global-plates-mobile-vit-v2-model', providers=['CPUExecutionProvider'])
    except:
        sys.exit(1)
    if len(sys.argv) > 1:
        image_path = sys.argv[1]
        result = process_plate_image(image_path, recognizer)
        if result:
            print(f"OCR_RESULT:{json.dumps(result)}")
    else:
        highest_conf_images = get_highest_confidence_images()
        if not highest_conf_images:
            sys.exit(1)
        for image_path in highest_conf_images:
            result = process_plate_image(image_path, recognizer)
            if result:
                print(f"OCR_RESULT:{json.dumps(result)}")